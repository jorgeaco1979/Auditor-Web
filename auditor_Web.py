import re
import io
import streamlit as st
import pandas as pd
import pypdf
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Auditor de Faturamento", page_icon="🏥", layout="wide")

# --- MOTOR DE EXTRAÇÃO (O mesmo que funcionou a 100%) ---
def extrair_dados_pdf_avancado(ficheiro_pdf, status_placeholder, barra_progresso):
    registros = []
    
    regex_credenciado = re.compile(r"CREDENCIADO:\s*(.*?)(?=\s*EXECUTOR:|\s*REGIME)", re.IGNORECASE)
    regex_valor = re.compile(r"V[lI]r\s*Pago:\s*([\d\.,]+)", re.IGNORECASE)
    regex_bloco_evento = re.compile(r"Cód\.\s*Evento:\s*([\d\.]+)\s*Qtd:\s*(\d+)\s*Evento:\s*(.*?)(?=V[lI]r\s*Pago:)", re.IGNORECASE)

    try:
        reader = pypdf.PdfReader(ficheiro_pdf)
        total_paginas = len(reader.pages)
        
        for num_pag in range(total_paginas):
            pagina = reader.pages[num_pag]
            texto_cru = pagina.extract_text()
            if not texto_cru:
                continue
            
            # Unificação total da página
            texto_unificado = texto_cru.replace('\n', ' ').replace('\r', ' ').replace('"', '')
            texto_unificado = re.sub(r'\s+', ' ', texto_unificado).strip()
            
            credenciados = regex_credenciado.findall(texto_unificado)
            valores = regex_valor.findall(texto_unificado)
            eventos_encontrados = list(regex_bloco_evento.finditer(texto_unificado))
            
            if eventos_encontrados:
                for i, match in enumerate(eventos_encontrados):
                    cod_evento = match.group(1).strip()
                    qtd = int(match.group(2).strip())
                    texto_procedimento = match.group(3).strip()
                    
                    # Faxina cirúrgica do texto
                    texto_procedimento = re.sub(r"Data Atend\s*\.:\s*[\d/]+", "", texto_procedimento, flags=re.IGNORECASE)
                    texto_procedimento = re.sub(r"Grau\s*:\s*PACOTE DESPESAS HOSPITALARES", "", texto_procedimento, flags=re.IGNORECASE)
                    texto_procedimento = re.sub(r"Grau\s*:\s*PACOTE", "", texto_procedimento, flags=re.IGNORECASE)
                    
                    if "(Código principal" in texto_procedimento:
                        texto_procedimento = texto_procedimento.split("(Código principal")[0]
                    if "(independente" in texto_procedimento:
                        texto_procedimento = texto_procedimento.split("(independente")[0]
                    
                    nome_final = texto_procedimento.replace(', ,', '').replace(',,', '').replace(' , ', '').strip()
                    nome_final = re.sub(r'^[,\s\.]+', '', nome_final)
                    nome_final = re.sub(r'[,\s\.]+$', '', nome_final)
                    nome_final = re.sub(r'\s+', ' ', nome_final).strip().upper()
                    
                    credenciado = credenciados[i].strip() if i < len(credenciados) else "Nao identificado"
                    credenciado = credenciado.strip(',').strip('.').strip().upper()
                    
                    if i < len(valores):
                        val_str = valores[i].replace('.', '').replace(',', '.')
                        try:
                            vlr_pago = float(val_str)
                        except ValueError:
                            vlr_pago = 0.0
                    else:
                        vlr_pago = 0.0
                    
                    registros.append({
                        "Codigo Evento": cod_evento,
                        "Procedimento": nome_final,
                        "Quantidade": qtd,
                        "Empresa/Credenciado": credenciado,
                        "Valor Pago (R$)": vlr_pago
                    })
            
            # Atualiza o ecrã na web
            progresso_atual = (num_pag + 1) / total_paginas
            barra_progresso.progress(progresso_atual)
            status_placeholder.text(f"A auditar página {num_pag + 1} de {total_paginas}...")
                
        return pd.DataFrame(registros)
    except Exception as e:
        return str(e)

def aplicar_estilo_excel(ws):
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = font_body
            if isinstance(cell.value, (float, int)):
                nome_coluna = str(ws.cell(row=1, column=cell.column).value).lower()
                if "valor" in nome_coluna or "financeiro" in nome_coluna or "(r$)" in nome_coluna or "pago" in nome_coluna:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center
            else:
                nome_coluna = str(ws.cell(row=1, column=cell.column).value).lower()
                if "codigo" in nome_coluna or "guia" in nome_coluna or "data" in nome_coluna:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format == 'R$ #,##0.00' and isinstance(cell.value, (float, int)):
                val_str = f"R$ {cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 80)

def gerar_excel_memoria(df, agrupado_codigo, agrupado_empresa, agrupado_cruzado):
    """Gera o Excel na memória (RAM) para download na web."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="1. Dados Brutos", index=False)
        agrupado_codigo.to_excel(writer, sheet_name="2. Total Geral por Codigo", index=False)
        agrupado_empresa.to_excel(writer, sheet_name="3. Ranking por Empresa")
        agrupado_cruzado.to_excel(writer, sheet_name="4. Empresa e Codigo", index=False)
        
        aplicar_estilo_excel(writer.sheets["1. Dados Brutos"])
        aplicar_estilo_excel(writer.sheets["2. Total Geral por Codigo"])
        aplicar_estilo_excel(writer.sheets["3. Ranking por Empresa"])
        aplicar_estilo_excel(writer.sheets["4. Empresa e Codigo"])
    
    return output.getvalue()

# --- INTERFACE WEB (FRONT-END) ---
st.title("📊 Auditor Médico de Faturação - PMDF")
st.markdown("Faça o upload do relatório da Benner (PDF) para extração automática e cruzamento de dados.")

ficheiro_carregado = st.file_uploader("Arraste e solte o ficheiro PDF aqui", type="pdf")

if ficheiro_carregado is not None:
    if st.button("🚀 Processar Relatório", type="primary"):
        status_texto = st.empty()
        barra = st.progress(0)
        
        # Extração
        df = extrair_dados_pdf_avancado(ficheiro_carregado, status_texto, barra)
        
        if isinstance(df, str):
            st.error(f"Erro ao ler o ficheiro: {df}")
        elif df.empty:
            st.warning("Nenhum dado encontrado no layout do PDF.")
        else:
            status_texto.text("A preparar cálculos executivos...")
            
            # Cálculos
            total_exames = df["Quantidade"].sum()
            valor_total = df["Valor Pago (R$)"].sum()
            
            agrupado_codigo = df.groupby(["Codigo Evento", "Procedimento"]).agg(
                Qtd_Total_Todas_Empresas=("Quantidade", "sum"),
                Valor_Total_Todas_Empresas=("Valor Pago (R$)", "sum")
            ).sort_values(by="Qtd_Total_Todas_Empresas", ascending=False).reset_index()
            
            agrupado_empresa = df.groupby("Empresa/Credenciado").agg(
                Qtd_Exames=("Quantidade", "sum"), 
                Financeiro_Total=("Valor Pago (R$)", "sum")
            ).sort_values(by="Financeiro_Total", ascending=False)
            
            agrupado_cruzado = df.groupby(["Empresa/Credenciado", "Codigo Evento", "Procedimento"]).agg(
                Qtd_Exames=("Quantidade", "sum"),
                Valor_Pago=("Valor Pago (R$)", "sum")
            ).sort_values(by=["Empresa/Credenciado", "Valor_Pago"], ascending=[True, False]).reset_index()
            
            # Conclui as barras
            barra.progress(100)
            status_texto.success("Análise concluída com sucesso!")
            
            # --- PAINEL EXECUTIVO (Métricas e Tabelas no Ecrã) ---
            st.markdown("---")
            col1, col2 = st.columns(2)
            col1.metric("📦 Volume Total de Exames", f"{total_exames:,.0f}".replace(',', '.'))
            col2.metric("💰 Impacto Financeiro Total", f"R$ {valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            
            # Tabs para as tabelas (Efeito visual fantástico)
            st.markdown("### Visualização de Dados")
            tab1, tab2, tab3 = st.tabs(["Cruzamento Detalhado", "Top Códigos Globais", "Ranking de Prestadores"])
            
            with tab1:
                st.dataframe(agrupado_cruzado, use_container_width=True)
            with tab2:
                st.dataframe(agrupado_codigo, use_container_width=True)
            with tab3:
                st.dataframe(agrupado_empresa, use_container_width=True)
            
            # --- BOTÃO DOWNLOAD EXCEL ---
            st.markdown("---")
            st.markdown("### Exportar Resultados")
            dados_excel = gerar_excel_memoria(df, agrupado_codigo, agrupado_empresa, agrupado_cruzado)
            
            st.download_button(
                label="📥 Descarregar Planilha Completa (Excel)",
                data=dados_excel,
                file_name="Auditoria_Benner_PMDF.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )